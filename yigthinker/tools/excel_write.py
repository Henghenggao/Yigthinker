"""excel_write — produce a formatted ``.xlsx`` file from a registered DataFrame,
with optional base-file modify mode, minimal named styles, and optional
native chart embedding.

This closes a gap observed in quick-260416-kyn: when the user asks for a
formatted P&L / report spreadsheet with section headers, subtotals, frozen
panes, and accounting number formats, the LLM has no first-class "write me a
styled xlsx" path and either (a) stuffs openpyxl source into df_transform or
(b) falls back to report_generate which only does bare header formatting.

Design choices (see ``.planning/quick/260416-kyn-.../260416-kyn-CONTEXT.md``):

- Output lands under ``~/.yigthinker/artifacts/<session_id>/`` (not the
  workspace), so every session has an isolated artifact bucket that the
  gateway's 7-day sweeper owns. Callers can override the filename, not the
  directory.
- ``base_file`` support: if set, the tool opens the base xlsx, adds/replaces
  ONE sheet, preserves every other sheet's formatting via openpyxl's native
  in-memory object graph, and writes the mutated workbook out under a
  deterministic new name. We never mutate the source file.
- Named styles are prefixed ``yt_`` to avoid collisions with user styles
  already present in the base workbook. Registering a named style that
  already exists raises ``ValueError`` in openpyxl ≥3.1, so we guard.
- ``row_style_map`` uses 0-based DataFrame row indices; row 0 in the map is
  the first data row (Excel row 2, since row 1 is the header).

Path safety reuses ``_safe_output_path`` from report_generate — the output
dir is under ``Path.home()`` but the attachments allowlist accepts explicit
registration so downstream adapters (e.g. Teams signed-URL delivery) can
both read from and serve the produced file.
"""
from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from pydantic import BaseModel, Field, field_validator

from yigthinker.session import SessionContext
from yigthinker.types import DryRunReceipt, ToolResult

logger = logging.getLogger(__name__)


# ── Plotly → openpyxl chart bridge ───────────────────────────────────────
#
# Translate the subset of Plotly chart types that openpyxl can natively
# render. openpyxl supports bar / line / pie / scatter / area out of the
# box; Plotly Express produces JSON whose x/y columns live in the axis
# titles (``layout.xaxis.title.text``), which is what we need to build a
# ``Reference`` into the already-written data range.
#
# Unsupported types (heatmap, scatter_3d, waterfall — the last NOT to be
# confused with openpyxl's line chart) return None so excel_write can
# skip the embed without erroring. Honesty over silent garbage: the tool
# surfaces ``embed_chart_skipped=True`` in its result so the LLM can
# narrate the degradation.

_PLOTLY_TO_OPENPYXL_CHART_KIND: dict[str, str] = {
    "bar": "bar",
    "line": "line",
    "scatter": "line",  # scatter-with-lines → openpyxl line
    "pie": "pie",
    "area": "line",     # approximate — openpyxl has no dedicated area chart
}


def _chart_spec_from_plotly_json(plotly_json: str) -> dict[str, Any] | None:
    """Extract a renderable chart spec from a Plotly figure JSON.

    Returns a dict with keys ``kind``, ``x_col``, ``y_col``, ``title`` if
    the chart is a type openpyxl can render. Returns ``None`` for
    unsupported types (heatmap, 3D scatter, etc.) or malformed JSON so
    callers can skip the embed cleanly.

    Column names come from the Plotly layout's axis-title ``text`` field,
    which Plotly Express populates automatically from the DataFrame
    column names (our ``chart_create`` tool always uses Plotly Express).
    """
    try:
        fig = json.loads(plotly_json)
    except (TypeError, json.JSONDecodeError):
        return None
    if not isinstance(fig, dict):
        return None

    data = fig.get("data") or []
    if not isinstance(data, list) or not data:
        return None
    trace = data[0]
    if not isinstance(trace, dict):
        return None

    trace_type = (trace.get("type") or "").lower()
    kind = _PLOTLY_TO_OPENPYXL_CHART_KIND.get(trace_type)
    if kind is None:
        return None

    layout = fig.get("layout") or {}

    # Title: can be a string or a dict with "text"
    title = ""
    layout_title = layout.get("title")
    if isinstance(layout_title, dict):
        title = str(layout_title.get("text") or "")
    elif isinstance(layout_title, str):
        title = layout_title

    # Pie: use labels + values (different Plotly field names)
    if kind == "pie":
        x_col = _extract_pie_label_col(trace, layout)
        y_col = _extract_pie_value_col(trace, layout)
    else:
        x_col = _extract_axis_title(layout, "xaxis") or trace.get("xaxis_title")
        y_col = _extract_axis_title(layout, "yaxis") or trace.get("yaxis_title")

    if not x_col or not y_col:
        return None

    return {
        "kind": kind,
        "x_col": str(x_col),
        "y_col": str(y_col),
        "title": title,
    }


def _extract_axis_title(layout: dict[str, Any], axis_key: str) -> str | None:
    axis = layout.get(axis_key) or {}
    if not isinstance(axis, dict):
        return None
    title = axis.get("title")
    if isinstance(title, dict):
        return str(title.get("text") or "") or None
    if isinstance(title, str):
        return title or None
    return None


def _extract_pie_label_col(trace: dict[str, Any], layout: dict[str, Any]) -> str | None:
    """Pie's category column — Plotly Express stores it in the trace
    hovertemplate as ``%{label}=...`` and in legend title."""
    # Easiest reliable source: Plotly Express puts the column name in
    # ``legendgroup`` or ``legend_title.text`` on the layout. As a final
    # fallback, we pull from the hovertemplate regex.
    legend = layout.get("legend") or {}
    if isinstance(legend, dict):
        lt = legend.get("title")
        if isinstance(lt, dict):
            txt = lt.get("text")
            if isinstance(txt, str) and txt:
                return txt
    # Parse from hovertemplate: "label=%{label}<br>..."  → the prefix before
    # "=" is often the column name.
    tpl = trace.get("hovertemplate") or ""
    if isinstance(tpl, str) and "=" in tpl:
        prefix = tpl.split("=", 1)[0].strip()
        if prefix and prefix.isidentifier():
            return prefix
    return None


def _extract_pie_value_col(trace: dict[str, Any], layout: dict[str, Any]) -> str | None:
    """Pie's numeric column — similar to label extraction but second
    occurrence in hovertemplate."""
    tpl = trace.get("hovertemplate") or ""
    if isinstance(tpl, str):
        # After the first "=%{...}", Plotly Express appends
        # "<br>value=%{value}". Pull the text between <br> and =.
        import re
        m = re.search(r"<br>([A-Za-z_][A-Za-z0-9_]*)=", tpl)
        if m:
            return m.group(1)
    return None


def _embed_native_chart(
    ws: Any,
    spec: dict[str, Any],
    df_headers: list[str],
    df_rows: int,
) -> bool:
    """Embed an openpyxl native chart into the given worksheet.

    The data is assumed to already be written: headers on row 1, data in
    rows 2..(df_rows+1), columns 1..len(df_headers). The chart is anchored
    two columns to the right of the last data column.

    Returns True on success, False if the chart spec's x_col / y_col
    aren't present in ``df_headers`` (schema mismatch). Caller decides
    whether to surface this as a skip in the tool result.
    """
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    kind = spec["kind"]
    x_col = spec["x_col"]
    y_col = spec["y_col"]

    if x_col not in df_headers or y_col not in df_headers:
        return False

    x_idx = df_headers.index(x_col) + 1   # 1-based Excel column
    y_idx = df_headers.index(y_col) + 1
    data_start_row = 1          # header row
    data_end_row = df_rows + 1  # last data row (inclusive)

    # y-data range includes the header so the chart's legend auto-picks up
    # the column name. openpyxl then uses titles_from_data=True.
    data_ref = Reference(
        ws,
        min_col=y_idx, max_col=y_idx,
        min_row=data_start_row, max_row=data_end_row,
    )
    cats_ref = Reference(
        ws,
        min_col=x_idx, max_col=x_idx,
        min_row=data_start_row + 1, max_row=data_end_row,
    )

    if kind == "bar":
        chart = BarChart()
        chart.type = "col"  # vertical bars (most common for finance)
    elif kind == "line":
        chart = LineChart()
    elif kind == "pie":
        chart = PieChart()
    else:  # pragma: no cover — map above should prevent this
        return False

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    if spec.get("title"):
        chart.title = spec["title"]

    # Anchor two columns to the right of the data, top-aligned with header
    from openpyxl.utils import get_column_letter
    anchor_col_letter = get_column_letter(len(df_headers) + 2)
    ws.add_chart(chart, f"{anchor_col_letter}1")
    return True

# ── Named-style registry ─────────────────────────────────────────────────
#
# Friendly names the LLM may use in ``row_style_map``. The rendered style
# name in the workbook is prefixed ``yt_`` to avoid colliding with user
# styles that might already exist in a base file.

_KNOWN_STYLE_NAMES: frozenset[str] = frozenset(
    {"section_header", "subtotal", "italic", "percent", "alt_row_fill"}
)

# Excel sheet-name rules (openpyxl enforces these but raises late — we
# reject eagerly so the LLM gets a clear schema-level error).
_FORBIDDEN_SHEET_CHARS: frozenset[str] = frozenset("[]:*?/\\")
_MAX_SHEET_NAME_LEN = 31

# Target file-size ceiling. openpyxl will happily build multi-GB workbooks
# and we do not want a runaway tool call to fill the disk.
_MAX_OUTPUT_BYTES = 50 * 1024 * 1024  # 50 MiB


class ExcelWriteInput(BaseModel):
    input_var: str = Field(
        description=(
            "Name of a registered DataFrame variable (see ``ctx.vars``). "
            "The DataFrame is written starting at cell A1 with headers on row 1."
        ),
    )
    sheet_name: str = Field(
        description=(
            "Target sheet name. Max 31 characters; must not contain any of "
            "``[ ] : * ? / \\``."
        ),
    )
    base_file: str | None = Field(
        default=None,
        description=(
            "Optional path to an existing ``.xlsx`` workbook to start from. "
            "Other sheets are preserved verbatim (formatting, formulas, "
            "merged cells). Must be registered in ``ctx.attachments`` or "
            "live inside the workspace."
        ),
    )
    overwrite_sheet: bool = Field(
        default=False,
        description=(
            "When ``base_file`` is set and ``sheet_name`` already exists in "
            "the base workbook: if False (default) the tool errors out; if "
            "True the existing sheet is dropped and replaced."
        ),
    )
    output_filename: str | None = Field(
        default=None,
        description=(
            "Basename (no directory) for the produced file. Defaults to "
            "``<base_stem>_<sheet_name>.xlsx`` in modify mode or "
            "``<sheet_name>.xlsx`` in new-file mode."
        ),
    )
    freeze_pane: str | None = Field(
        default=None,
        description=(
            "Cell reference (e.g. ``'B2'``) marking the top-left cell of "
            "the scrollable region. Use ``'A2'`` to freeze the header row."
        ),
    )
    number_format: dict[str, str] | None = Field(
        default=None,
        description=(
            "Map column-name → openpyxl number-format code, e.g. "
            "``{'Revenue': '#,##0;[Red]-#,##0', 'Margin': '0.0%'}``."
        ),
    )
    row_style_map: dict[int, str] | None = Field(
        default=None,
        description=(
            "Map 0-based DataFrame row index → named style. Valid names: "
            f"{sorted(_KNOWN_STYLE_NAMES)}."
        ),
    )
    summary: str | None = Field(
        default=None,
        description="Optional one-liner shown on the IM card.",
    )
    embed_chart: str | None = Field(
        default=None,
        description=(
            "Optional: name of a chart variable in ``ctx.vars`` (produced by "
            "chart_create) to embed as an openpyxl native chart in the "
            "output workbook. Supported chart types: bar, line, pie. The "
            "chart's x and y columns must match columns in the DataFrame "
            "being written. Anchored two columns right of the data. If the "
            "chart is an unsupported type (heatmap / scatter_3d / "
            "waterfall) or its columns don't match the data schema, the "
            "embed is skipped with ``embed_chart_skipped=True`` in the "
            "result — the xlsx is still produced, just without the chart."
        ),
    )

    @field_validator("sheet_name")
    @classmethod
    def _check_sheet_name(cls, v: str) -> str:
        if not v or not v.strip():
            raise ValueError("sheet_name must not be empty")
        if len(v) > _MAX_SHEET_NAME_LEN:
            raise ValueError(
                f"sheet_name too long ({len(v)} chars). Excel caps at "
                f"{_MAX_SHEET_NAME_LEN}."
            )
        bad = sorted(ch for ch in v if ch in _FORBIDDEN_SHEET_CHARS)
        if bad:
            raise ValueError(
                f"sheet_name contains forbidden character(s) {bad!r}. "
                "Excel disallows any of [ ] : * ? / \\."
            )
        return v


class ExcelWriteTool:
    name = "excel_write"
    description = (
        "Write a formatted ``.xlsx`` file from a registered DataFrame with "
        "named styles (section_header / subtotal / italic / percent / "
        "alt_row_fill), per-column number formats, and optional frozen "
        "panes. Supports ``base_file`` modify mode: open an existing "
        "workbook, add or replace ONE sheet, and preserve every other "
        "sheet's formatting. Use THIS instead of df_transform when the "
        "user asks for a formatted xlsx — df_transform cannot style "
        "cells and report_generate only applies bare header formatting. "
        "Output lands under the session artifacts dir and is auto-"
        "registered in ``ctx.attachments`` for IM delivery."
    )
    input_schema = ExcelWriteInput

    async def execute(
        self, input: ExcelWriteInput, ctx: SessionContext,
    ) -> ToolResult:
        if ctx.dry_run:
            return ToolResult(
                tool_use_id="",
                content=DryRunReceipt(
                    tool_name=self.name,
                    summary=(
                        f"Would write Excel sheet '{input.sheet_name}' "
                        f"from {input.input_var}"
                        + (f" into base_file {input.base_file}"
                           if input.base_file else "")
                    ),
                    details={"input": input.model_dump()},
                ),
            )

        # ── 1. Resolve the DataFrame ─────────────────────────────────
        try:
            df = ctx.vars.get(input.input_var)
        except KeyError as exc:
            return ToolResult(tool_use_id="", content=str(exc), is_error=True)

        try:
            import pandas as pd  # local — pandas is a core dep
        except ImportError as exc:  # pragma: no cover — core dep
            return ToolResult(
                tool_use_id="", content=f"pandas not available: {exc}",
                is_error=True,
            )
        if not isinstance(df, pd.DataFrame):
            return ToolResult(
                tool_use_id="",
                content=(
                    f"Variable '{input.input_var}' is a {type(df).__name__}, "
                    "not a DataFrame. excel_write expects a tabular input."
                ),
                is_error=True,
            )

        # ── 2. Validate row_style_map friendly names early ───────────
        if input.row_style_map:
            unknown = sorted(
                set(input.row_style_map.values()) - _KNOWN_STYLE_NAMES
            )
            if unknown:
                return ToolResult(
                    tool_use_id="",
                    content=(
                        f"Unknown style name(s) {unknown!r}. "
                        f"Valid: {sorted(_KNOWN_STYLE_NAMES)}."
                    ),
                    is_error=True,
                )

        # ── 3. Validate & open base_file (if any) ────────────────────
        base_wb = None
        base_stem: str | None = None
        if input.base_file is not None:
            # Import locally so test_rejects_invalid_sheet_name_* don't pay
            # the openpyxl import cost.
            from yigthinker.tools.reports.report_generate import (
                _safe_output_path,
            )
            base_resolved, err = _safe_output_path(
                input.base_file, ctx.settings, attachments=ctx.attachments,
            )
            if err:
                return ToolResult(tool_use_id="", content=err, is_error=True)
            if not base_resolved.exists():
                return ToolResult(
                    tool_use_id="",
                    content=f"base_file does not exist: {base_resolved}",
                    is_error=True,
                )
            if base_resolved.suffix.lower() == ".xlsm":
                return ToolResult(
                    tool_use_id="",
                    content=(
                        "base_file is an .xlsm (macro-enabled) workbook. "
                        "excel_write refuses to open macro files — re-save "
                        "as plain .xlsx first."
                    ),
                    is_error=True,
                )
            if base_resolved.suffix.lower() != ".xlsx":
                return ToolResult(
                    tool_use_id="",
                    content=(
                        f"base_file must be .xlsx, got "
                        f"{base_resolved.suffix!r}."
                    ),
                    is_error=True,
                )

            import openpyxl
            try:
                base_wb = openpyxl.load_workbook(filename=str(base_resolved))
            except Exception as exc:  # pragma: no cover — openpyxl raises many
                return ToolResult(
                    tool_use_id="",
                    content=f"Failed to open base_file: {exc}",
                    is_error=True,
                )

            # Conflict check against existing sheet
            if input.sheet_name in base_wb.sheetnames and not input.overwrite_sheet:
                return ToolResult(
                    tool_use_id="",
                    content=(
                        f"Sheet '{input.sheet_name}' already exists in "
                        f"{base_resolved.name}. Pass overwrite_sheet=True "
                        "to replace it, or pick a different sheet_name."
                    ),
                    is_error=True,
                )
            base_stem = base_resolved.stem

        # ── 4. Resolve output path ───────────────────────────────────
        out_dir = _artifacts_dir(ctx)
        out_dir.mkdir(parents=True, exist_ok=True)

        if input.output_filename:
            # Basename only — no slashes, no parent escapes.
            safe_name = Path(input.output_filename).name
            if not safe_name.lower().endswith(".xlsx"):
                safe_name = f"{safe_name}.xlsx"
        elif base_stem is not None:
            safe_name = f"{base_stem}_{input.sheet_name}.xlsx"
        else:
            safe_name = f"{input.sheet_name}.xlsx"

        out_path = (out_dir / safe_name).resolve()

        # ── 5. Build / mutate the workbook ───────────────────────────
        import openpyxl

        if base_wb is not None:
            wb = base_wb
            if input.sheet_name in wb.sheetnames:
                # overwrite_sheet is True (we short-circuited earlier if not).
                del wb[input.sheet_name]
            ws = wb.create_sheet(title=input.sheet_name)
        else:
            wb = openpyxl.Workbook()
            default_ws = wb.active
            default_ws.title = input.sheet_name
            ws = default_ws

        # Register named styles (idempotent — re-registering raises ValueError
        # in openpyxl ≥3.1 so we guard by checking wb.named_styles).
        _register_named_styles(wb)

        # ── 6. Write the DataFrame (headers row 1, data row 2+) ──────
        headers = [str(c) for c in df.columns]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, (_, row) in enumerate(df.iterrows()):
            excel_row = row_idx + 2  # +1 for 1-based, +1 for header row
            for col_idx, value in enumerate(row.tolist(), start=1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.value = _coerce_value(value)

        # ── 7. Apply per-column number formats ───────────────────────
        if input.number_format:
            header_to_col = {h: i + 1 for i, h in enumerate(headers)}
            data_row_count = len(df.index)
            for col_name, fmt in input.number_format.items():
                col_idx = header_to_col.get(col_name)
                if col_idx is None:
                    continue  # silent — the LLM sometimes names optional cols
                for r in range(2, data_row_count + 2):
                    ws.cell(row=r, column=col_idx).number_format = fmt

        # ── 8. Apply row styles ──────────────────────────────────────
        if input.row_style_map:
            col_count = max(1, len(headers))
            for df_row, friendly in input.row_style_map.items():
                excel_row = df_row + 2  # 0-based DF → Excel data row
                style_name = f"yt_{friendly}"
                if style_name not in wb.named_styles:
                    continue  # defensive — validated above
                for col_idx in range(1, col_count + 1):
                    ws.cell(row=excel_row, column=col_idx).style = style_name

        # ── 9. Freeze panes ──────────────────────────────────────────
        if input.freeze_pane:
            ws.freeze_panes = input.freeze_pane

        # ── 9b. Embed native chart (optional, 2026-04-18 UAT follow-up) ─
        embed_chart_skipped = False
        embed_chart_skip_reason: str | None = None
        if input.embed_chart:
            chart_json_str = None
            try:
                raw = ctx.vars.get(input.embed_chart)
            except KeyError:
                # List available charts so the LLM can self-correct
                available_charts = sorted(
                    info.name
                    for info in ctx.vars.list()
                    if getattr(info, "var_type", "") == "chart"
                )
                charts_listing = (
                    ", ".join(available_charts)
                    if available_charts
                    else "(none — call chart_create first)"
                )
                return ToolResult(
                    tool_use_id="",
                    content=(
                        f"embed_chart: chart '{input.embed_chart}' not found in "
                        f"ctx.vars. Available chart names: {charts_listing}"
                    ),
                    is_error=True,
                )
            if isinstance(raw, str):
                chart_json_str = raw
            else:
                embed_chart_skipped = True
                embed_chart_skip_reason = (
                    f"registered variable '{input.embed_chart}' is a "
                    f"{type(raw).__name__}, not a chart JSON string. "
                    f"Expected output of chart_create."
                )
            if chart_json_str is not None:
                spec = _chart_spec_from_plotly_json(chart_json_str)
                if spec is None:
                    embed_chart_skipped = True
                    embed_chart_skip_reason = (
                        "chart type is not supported for native openpyxl "
                        "embedding (supported: bar, line, pie, area, "
                        "scatter). The xlsx was still produced without "
                        "the chart."
                    )
                else:
                    ok = _embed_native_chart(
                        ws, spec, headers, df_rows=len(df.index),
                    )
                    if not ok:
                        embed_chart_skipped = True
                        embed_chart_skip_reason = (
                            f"chart columns (x='{spec['x_col']}', "
                            f"y='{spec['y_col']}') do not match any columns "
                            f"in the DataFrame being written (columns: "
                            f"{headers}). Re-run chart_create on the same "
                            f"DataFrame to produce a matching chart."
                        )

        # ── 10. Save ─────────────────────────────────────────────────
        try:
            wb.save(str(out_path))
        except Exception as exc:  # pragma: no cover — disk / perm errors
            return ToolResult(
                tool_use_id="",
                content=f"Failed to save {out_path}: {exc}",
                is_error=True,
            )

        try:
            size_bytes = out_path.stat().st_size
        except OSError:
            size_bytes = 0
        if size_bytes > _MAX_OUTPUT_BYTES:
            # Unlink the oversized file to avoid leaking disk — the LLM
            # should either slim the DF or use report_generate/csv.
            out_path.unlink(missing_ok=True)
            return ToolResult(
                tool_use_id="",
                content=(
                    f"Output exceeded {_MAX_OUTPUT_BYTES:,}-byte cap "
                    f"({size_bytes:,} bytes). Trim the DataFrame or split."
                ),
                is_error=True,
            )

        # Register so IM adapters can surface / serve the file.
        ctx.attachments.add(out_path)

        content: dict[str, Any] = {
            "kind": "file",
            "path": str(out_path),
            "filename": out_path.name,
            "bytes": size_bytes,
            "summary": input.summary,
            "mime_type": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        }
        # Honest degradation flag so the LLM can narrate "Excel was saved
        # but the chart embed was skipped because <reason>" rather than
        # silently implying the chart is in the xlsx.
        if input.embed_chart is not None:
            content["embed_chart"] = input.embed_chart
            content["embed_chart_skipped"] = embed_chart_skipped
            if embed_chart_skip_reason is not None:
                content["embed_chart_skip_reason"] = embed_chart_skip_reason
        return ToolResult(
            tool_use_id="",
            content=content,
        )


# ── Module-private helpers ───────────────────────────────────────────────


def _artifacts_dir(ctx: SessionContext) -> Path:
    """Return ``~/.yigthinker/artifacts/<session_id>/`` for this session."""
    return Path.home() / ".yigthinker" / "artifacts" / ctx.session_id


def _coerce_value(value: Any) -> Any:
    """Convert pandas sentinel values to something openpyxl accepts.

    openpyxl rejects ``pd.NaT`` and silently mis-writes numpy scalars that
    pandas returns from ``DataFrame.iterrows()``. Round-trip via Python
    scalars for safety.
    """
    try:
        import pandas as pd  # noqa: WPS433
        if pd.isna(value):
            return None
    except Exception:
        pass
    # numpy scalar → Python scalar
    item = getattr(value, "item", None)
    if callable(item):
        try:
            return item()
        except Exception:
            return value
    return value


def _register_named_styles(wb: Any) -> None:
    """Register the minimal ``yt_*`` named style set on the workbook.

    openpyxl raises ``ValueError("Style ... exists already")`` if you add
    the same NamedStyle twice — which happens when the base workbook was
    itself produced by a prior excel_write call. We guard by checking
    ``wb.named_styles`` (the list-of-names view).
    """
    from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill

    existing = set(wb.named_styles)

    styles: dict[str, NamedStyle] = {}

    # section_header — bold, light-grey fill, left aligned
    styles["yt_section_header"] = NamedStyle(
        name="yt_section_header",
        font=Font(bold=True),
        fill=PatternFill("solid", fgColor="E7E6E6"),
        alignment=Alignment(horizontal="left", vertical="center"),
    )

    # subtotal — bold, top-border feel via font + italic off
    styles["yt_subtotal"] = NamedStyle(
        name="yt_subtotal",
        font=Font(bold=True),
    )

    # italic — for footnote / commentary rows
    styles["yt_italic"] = NamedStyle(
        name="yt_italic",
        font=Font(italic=True),
    )

    # percent — number_format for percentage cells
    styles["yt_percent"] = NamedStyle(
        name="yt_percent",
        number_format="0.0%",
    )

    # alt_row_fill — subtle zebra striping
    styles["yt_alt_row_fill"] = NamedStyle(
        name="yt_alt_row_fill",
        fill=PatternFill("solid", fgColor="F2F2F2"),
    )

    for name, ns in styles.items():
        if name in existing:
            continue
        try:
            wb.add_named_style(ns)
        except ValueError:
            # Race: another code path added the style between our check
            # and the add. Safe to ignore.
            continue
