"""Tests for the excel_write tool (quick-260416-kyn).

Covers writing a new xlsx from a DataFrame, modifying an existing base file
(add/replace sheet with preserved formatting on other sheets), named styles,
freeze panes, per-column number_format, and row_style_map. Uses an in-memory
``BytesIO`` round-trip for post-write inspection so tests are Windows-tmpdir
clean (per CONTEXT.md Claude's Discretion §Test fixtures).
"""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from yigthinker.session import SessionContext
from yigthinker.tools.excel_write import (
    _KNOWN_STYLE_NAMES,
    ExcelWriteTool,
)


def _make_ctx(tmp_path: Path, monkeypatch) -> SessionContext:
    """Build a SessionContext whose ~/.yigthinker/artifacts/<sid>/ lands under tmp_path.

    We monkeypatch ``Path.home`` only on the ``yigthinker.tools.excel_write``
    module's imported symbol so the tool writes into the test tmpdir instead of
    the real user home.
    """
    # Redirect Path.home() -> tmp_path in the excel_write module.
    import yigthinker.tools.excel_write as ew_mod
    monkeypatch.setattr(ew_mod.Path, "home", classmethod(lambda cls: tmp_path))
    return SessionContext(settings={"workspace_dir": str(tmp_path)})


def _load_bytes(path: Path) -> openpyxl.Workbook:
    """In-memory round-trip — avoids Windows tmpdir flakiness with file handles."""
    return openpyxl.load_workbook(filename=io.BytesIO(path.read_bytes()))


def _simple_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Account": ["Revenue", "Direct Costs", "Gross Profit"],
            "Jan": [100, 40, 60],
            "Feb": [120, 50, 70],
        }
    )


# ── Pydantic validation tests ────────────────────────────────────────────


def test_rejects_invalid_sheet_name_too_long():
    tool = ExcelWriteTool()
    with pytest.raises(Exception):  # pydantic ValidationError
        tool.input_schema(input_var="df1", sheet_name="A" * 32)


def test_rejects_invalid_sheet_name_forbidden_chars():
    tool = ExcelWriteTool()
    for bad in ["has/slash", "has\\slash", "has*star", "has?", "has[bracket]", "has:colon"]:
        with pytest.raises(Exception):
            tool.input_schema(input_var="df1", sheet_name=bad)


# ── Happy-path / new-file creation ────────────────────────────────────────


async def test_writes_new_xlsx(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)
    ctx.vars.set("pl", _simple_df())

    result = await tool.execute(
        tool.input_schema(input_var="pl", sheet_name="P&L"),
        ctx,
    )
    assert not result.is_error, result.content
    assert isinstance(result.content, dict)
    assert result.content["kind"] == "file"
    assert result.content["mime_type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    out_path = Path(result.content["path"])
    assert out_path.exists()
    # Registered in attachments for downstream adapters
    assert out_path.resolve() in ctx.attachments

    wb = _load_bytes(out_path)
    assert "P&L" in wb.sheetnames
    ws = wb["P&L"]
    # Headers in row 1, data starting row 2
    headers = [cell.value for cell in ws[1]]
    assert headers == ["Account", "Jan", "Feb"]
    assert ws.cell(row=2, column=1).value == "Revenue"
    assert ws.cell(row=4, column=3).value == 70
    # Named styles registered on the new workbook
    for friendly in _KNOWN_STYLE_NAMES:
        assert f"yt_{friendly}" in wb.named_styles


# ── Base-file modify mode ─────────────────────────────────────────────────


async def test_modifies_existing_base_file(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)

    # Build a pre-existing base workbook with a "Forecast" sheet.
    base_path = (tmp_path / "FCST_2+10 10.xlsx").resolve()
    base_wb = openpyxl.Workbook()
    base_ws = base_wb.active
    base_ws.title = "Forecast"
    base_ws["A1"] = "KeepMe"
    base_ws["A1"].font = openpyxl.styles.Font(bold=True)
    base_wb.save(base_path)
    # Register in attachments so _safe_output_path accepts it
    ctx.attachments.add(base_path)

    ctx.vars.set("pl", _simple_df())
    result = await tool.execute(
        tool.input_schema(
            input_var="pl",
            sheet_name="P&L",
            base_file=str(base_path),
            summary="Added P&L sheet",
        ),
        ctx,
    )
    assert not result.is_error, result.content
    out_path = Path(result.content["path"])
    # Default naming: <base_stem>_<sheet_name>.xlsx under artifacts/<sid>/
    assert out_path.name == "FCST_2+10 10_P&L.xlsx"
    assert out_path.parent.name == ctx.session_id
    assert out_path.parent.parent == (tmp_path / ".yigthinker" / "artifacts")

    wb = _load_bytes(out_path)
    # Both sheets preserved
    assert "Forecast" in wb.sheetnames
    assert "P&L" in wb.sheetnames
    # Original sheet untouched
    forecast = wb["Forecast"]
    assert forecast["A1"].value == "KeepMe"
    assert forecast["A1"].font.bold is True
    # New sheet has our data
    pl = wb["P&L"]
    assert pl["A1"].value == "Account"


async def test_overwrite_sheet_true(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)

    base_path = (tmp_path / "book.xlsx").resolve()
    base_wb = openpyxl.Workbook()
    base_wb.active.title = "P&L"
    base_wb["P&L"]["A1"] = "OldValue"
    base_wb.save(base_path)
    ctx.attachments.add(base_path)

    ctx.vars.set("pl", _simple_df())
    result = await tool.execute(
        tool.input_schema(
            input_var="pl",
            sheet_name="P&L",
            base_file=str(base_path),
            overwrite_sheet=True,
        ),
        ctx,
    )
    assert not result.is_error, result.content
    wb = _load_bytes(Path(result.content["path"]))
    assert wb["P&L"]["A1"].value == "Account"  # replaced, not OldValue


async def test_overwrite_sheet_false_errors(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)

    base_path = (tmp_path / "book.xlsx").resolve()
    base_wb = openpyxl.Workbook()
    base_wb.active.title = "P&L"
    base_wb.save(base_path)
    ctx.attachments.add(base_path)

    ctx.vars.set("pl", _simple_df())
    result = await tool.execute(
        tool.input_schema(
            input_var="pl",
            sheet_name="P&L",
            base_file=str(base_path),
            overwrite_sheet=False,
        ),
        ctx,
    )
    assert result.is_error
    assert "already exists" in str(result.content).lower() or "overwrite" in str(result.content).lower()


# ── Formatting application ────────────────────────────────────────────────


async def test_freeze_pane_applied(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)
    ctx.vars.set("pl", _simple_df())

    result = await tool.execute(
        tool.input_schema(input_var="pl", sheet_name="P&L", freeze_pane="B2"),
        ctx,
    )
    assert not result.is_error, result.content
    wb = _load_bytes(Path(result.content["path"]))
    assert wb["P&L"].freeze_panes == "B2"


async def test_number_format_applied(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)
    ctx.vars.set("pl", _simple_df())

    fmt = "#,##0;[Red]-#,##0"
    result = await tool.execute(
        tool.input_schema(
            input_var="pl",
            sheet_name="P&L",
            number_format={"Jan": fmt},
        ),
        ctx,
    )
    assert not result.is_error, result.content
    wb = _load_bytes(Path(result.content["path"]))
    ws = wb["P&L"]
    # Data rows start at row 2 (row 1 is header). Jan is col 2.
    for r in range(2, 5):
        assert ws.cell(row=r, column=2).number_format == fmt


async def test_row_style_map_applied(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)
    ctx.vars.set("pl", _simple_df())

    result = await tool.execute(
        tool.input_schema(
            input_var="pl",
            sheet_name="P&L",
            row_style_map={0: "section_header", 2: "subtotal"},
        ),
        ctx,
    )
    assert not result.is_error, result.content
    wb = _load_bytes(Path(result.content["path"]))
    ws = wb["P&L"]
    # DataFrame row 0 = Excel row 2; row 2 = Excel row 4
    # Check the row has the named style on at least one cell (all in a styled
    # row are set by the tool)
    assert ws.cell(row=2, column=1).style == "yt_section_header"
    assert ws.cell(row=4, column=1).style == "yt_subtotal"


async def test_duplicate_named_style_safe(tmp_path, monkeypatch):
    """When the base file already has yt_section_header registered, re-registering
    must NOT raise — we guard against NamedStyle duplicate-add ValueError.
    """
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)

    base_path = (tmp_path / "pre_styled.xlsx").resolve()
    base_wb = openpyxl.Workbook()
    from openpyxl.styles import NamedStyle
    ns = NamedStyle(name="yt_section_header")
    base_wb.add_named_style(ns)
    base_wb.save(base_path)
    ctx.attachments.add(base_path)

    ctx.vars.set("pl", _simple_df())
    result = await tool.execute(
        tool.input_schema(
            input_var="pl",
            sheet_name="P&L",
            base_file=str(base_path),
        ),
        ctx,
    )
    assert not result.is_error, result.content


# ── Error paths ───────────────────────────────────────────────────────────


async def test_rejects_base_file_outside_allowlist(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)
    ctx.vars.set("pl", _simple_df())

    # base_file exists on disk but is NOT registered in ctx.attachments and is
    # outside workspace_dir
    outside = tmp_path.parent / "totally_outside.xlsx"
    base_wb = openpyxl.Workbook()
    base_wb.save(outside)
    try:
        result = await tool.execute(
            tool.input_schema(
                input_var="pl", sheet_name="P&L", base_file=str(outside),
            ),
            ctx,
        )
        assert result.is_error
        assert "outside" in str(result.content).lower() or "access" in str(result.content).lower()
    finally:
        outside.unlink(missing_ok=True)


async def test_rejects_unknown_style_name(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)
    ctx.vars.set("pl", _simple_df())

    result = await tool.execute(
        tool.input_schema(
            input_var="pl",
            sheet_name="P&L",
            row_style_map={0: "mauve"},
        ),
        ctx,
    )
    assert result.is_error
    # Error lists the valid names so the LLM can fix it
    msg = str(result.content).lower()
    assert "mauve" in msg
    assert "section_header" in msg


async def test_rejects_missing_var(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)
    # No variable registered

    result = await tool.execute(
        tool.input_schema(input_var="missing", sheet_name="P&L"),
        ctx,
    )
    assert result.is_error
    assert "missing" in str(result.content).lower() or "not found" in str(result.content).lower()


async def test_rejects_xlsm_base_file(tmp_path, monkeypatch):
    tool = ExcelWriteTool()
    ctx = _make_ctx(tmp_path, monkeypatch)
    ctx.vars.set("pl", _simple_df())

    # Path doesn't need to exist — the suffix check happens after _safe_output_path
    fake_xlsm = (tmp_path / "macros.xlsm").resolve()
    fake_xlsm.write_bytes(b"not really an xlsm but suffix check is lexical")
    ctx.attachments.add(fake_xlsm)

    result = await tool.execute(
        tool.input_schema(
            input_var="pl", sheet_name="P&L", base_file=str(fake_xlsm),
        ),
        ctx,
    )
    assert result.is_error
    assert "xlsm" in str(result.content).lower()


# ── artifacts.py passthrough ──────────────────────────────────────────────


def test_mime_type_passthrough():
    """structured_artifact_from_tool_result must preserve mime_type when present
    in the file-kind payload (excel_write sets it; artifact_write does not).
    """
    from yigthinker.channels.artifacts import structured_artifact_from_tool_result

    raw = {
        "kind": "file",
        "path": "/tmp/x.xlsx",
        "filename": "x.xlsx",
        "bytes": 10,
        "summary": None,
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    artifact = structured_artifact_from_tool_result(raw)
    assert artifact is not None
    assert artifact["mime_type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Legacy payload (no mime_type) still works — no KeyError
    legacy = {
        "kind": "file",
        "path": "/tmp/a.py",
        "filename": "a.py",
        "bytes": 5,
        "summary": None,
    }
    legacy_artifact = structured_artifact_from_tool_result(legacy)
    assert legacy_artifact is not None
    assert legacy_artifact.get("mime_type") is None


# ── Registry wiring ───────────────────────────────────────────────────────


def test_registry_registration():
    """build_tool_registry must include excel_write (quick-260416-kyn)."""
    from yigthinker.registry_factory import build_tool_registry
    from yigthinker.tools.sql.connection import ConnectionPool

    registry = build_tool_registry(ConnectionPool())
    assert "excel_write" in registry.names()


def test_tool_description_nudges_away_from_df_transform():
    """LLM nudging — research open risk #8. excel_write must redirect the LLM
    away from df_transform for formatted xlsx output.
    """
    desc = ExcelWriteTool.description
    assert "df_transform" in desc or "formatted" in desc.lower()
    # Must still describe core capability
    assert "xlsx" in desc.lower() or "excel" in desc.lower()
